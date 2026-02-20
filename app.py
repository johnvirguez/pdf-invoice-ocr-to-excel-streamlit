import io
import re
from dataclasses import dataclass
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from pypdf import PdfReader

from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# =========================
# PAGE + STYLE (UX)
# =========================
st.set_page_config(page_title="SED | Facturas → Excel (Finanzas)", layout="wide")

st.markdown(
    """
    <style>
      .file-scroll {
        max-height: 220px;
        overflow-y: auto;
        padding: 8px 10px;
        border: 1px solid rgba(49, 51, 63, 0.2);
        border-radius: 10px;
        background: rgba(250, 250, 250, 0.6);
      }
      .table-scroll {
        max-height: 520px;
        overflow-y: auto;
        border: 1px solid rgba(49, 51, 63, 0.2);
        border-radius: 10px;
        padding: 6px;
        background: white;
      }
      div[data-testid="stDataFrame"] { margin-top: 0.25rem; }
    </style>
    """,
    unsafe_allow_html=True,
)

APP_TITLE = "📄➡️📊 SED | Facturas PDF → Excel (Contabilidad/Finanzas)"
APP_SUBTITLE = (
    "Carga PDFs → Procesa → Selecciona una factura para ver el detalle (zoom) → Descarga Excel final. "
    "Excel: agrupado por factura, sin cuadrícula y fuente Century Gothic 10."
)

st.title(APP_TITLE)
st.caption(APP_SUBTITLE)


# =========================
# OUTPUT COLUMNS
# =========================
FIN_COLS = [
    "Documento",
    "Pais",
    "Tipo_Documento",
    "Proveedor_Razon_Social",
    "Proveedor_Id_Tributaria",
    "Cliente_Razon_Social",
    "Cliente_Id_Tributaria",
    "Prefijo",
    "Factura_Numero",
    "Consecutivo",
    "Fecha_Emision",
    "Condicion_Venta",
    "Forma_Pago",
    "Medio_Pago",
    "Moneda",
    "Simbolo_Moneda",
    "Subtotal",
    "Impuesto_IVA",
    "Total_Factura",
    "OC",
    "CUFE",
    "Resolucion_DIAN",
    "QR_o_Codigo",
    "Clave_Numerica",
    "Codigo_Unico_Consulta",
    "Anticipo",
    "Saldo",
    "Costo_Factura_Marcado",
    "Probable_Escaneado",
    "Metodo_Extraccion",
    "Error",
]

LINE_COLS = [
    "Factura_Numero",
    "Documento",
    "Linea",
    "Codigo_Item",
    "Descripcion",
    "Cantidad",
    "Unidad",
    "Precio_Unitario",
    "Descuento",
    "Subtotal_Linea",
    "Impuesto_Linea",
    "Total_Linea",
    "Moneda",
    "Pais",
    "Marca_Costo",
    "Cuenta_Costo",
    "Descripcion_Raw",
]


# =========================
# DATA STRUCTURES
# =========================
@dataclass
class FinanceInvoice:
    Documento: str
    Pais: str = ""
    Tipo_Documento: str = "Factura"

    Proveedor_Razon_Social: str = ""
    Proveedor_Id_Tributaria: str = ""
    Cliente_Razon_Social: str = ""
    Cliente_Id_Tributaria: str = ""

    Prefijo: str = ""
    Factura_Numero: str = ""
    Consecutivo: str = ""
    Fecha_Emision: str = ""

    Condicion_Venta: str = ""
    Forma_Pago: str = ""
    Medio_Pago: str = ""

    Moneda: str = ""
    Simbolo_Moneda: str = ""
    Subtotal: Optional[float] = None
    Impuesto_IVA: Optional[float] = None
    Total_Factura: Optional[float] = None

    OC: str = ""
    CUFE: str = ""
    Resolucion_DIAN: str = ""
    QR_o_Codigo: str = ""

    Clave_Numerica: str = ""
    Codigo_Unico_Consulta: str = ""

    Anticipo: Optional[float] = None
    Saldo: Optional[float] = None

    Costo_Factura_Marcado: str = ""
    Probable_Escaneado: str = ""
    Metodo_Extraccion: str = ""
    Error: str = ""


# =========================
# TEXT UTILITIES
# =========================
def normalize_text(t: str) -> str:
    if t is None:
        return ""
    t = t.replace("\u00a0", " ")
    t = re.sub(r"[ \t]+", " ", t)
    t = re.sub(r"\n{3,}", "\n\n", t)
    return t.strip()


def extract_text_pypdf(pdf_bytes: bytes) -> str:
    reader = PdfReader(io.BytesIO(pdf_bytes))
    parts: List[str] = []
    for page in reader.pages:
        parts.append(page.extract_text() or "")
    return normalize_text("\n".join(parts))


def looks_scanned(text: str) -> bool:
    return len((text or "").strip()) < 50


def safe_group(m: Optional[re.Match], idx: int = 1) -> str:
    if not m:
        return ""
    try:
        val = m.group(idx) if m.lastindex and idx <= m.lastindex else m.group(0)
    except Exception:
        val = m.group(0) if m else ""
    return (val or "").strip()


def find_first(patterns: List[str], text: str, flags=re.IGNORECASE) -> str:
    if not text:
        return ""
    for p in patterns:
        m = re.search(p, text, flags)
        if m:
            if m.lastindex and m.lastindex >= 1:
                return safe_group(m, 1)
            return safe_group(m, 0)
    return ""


def parse_number_latam(s: str) -> Optional[float]:
    if not s:
        return None
    raw = s.strip()
    raw = re.sub(r"[^\d,.\-]", "", raw)
    if not raw:
        return None

    last_comma = raw.rfind(",")
    last_dot = raw.rfind(".")

    if last_comma > last_dot:
        raw = raw.replace(".", "")
        raw = raw.replace(",", ".")
    else:
        raw = raw.replace(",", "")

    try:
        return float(raw)
    except:
        return None


def detect_currency(text: str) -> Tuple[str, str]:
    t = text or ""
    if " CRC" in t or "Moneda: CRC" in t or "Código Moneda........ CRC" in t or "¢" in t:
        return "CRC", "¢"
    if " COP" in t or "pesos" in t.lower() or "Bogotá - Colombia" in t or "$" in t:
        return "COP", "$"
    return "", ""


def lines(text: str) -> List[str]:
    return [ln.strip() for ln in (text or "").splitlines() if (ln or "").strip()]


# =========================
# FORMAT DETECTORS
# =========================
def is_forlan_co(text: str) -> bool:
    t = (text or "").upper()
    return "FERRETERIA FORLAN" in t and "FACTURA ELECTR" in t and "TOTAL A PAGAR" in t


def is_navatec_cr(text: str) -> bool:
    t = (text or "").upper()
    return ("FACTURA ELECTRÓNICA N°" in t or "FACTURA ELECTRONICA N°" in t) and "FACTURAELECTRONICA.CR" in t


def is_tribu_cr_hacienda(text: str) -> bool:
    t = (text or "").upper()
    return "WWW.HACIENDA.GO.CR" in t and "TRIBU-CR" in t and "COMPROBANTE" in t


def is_ciclo_huracan(text: str) -> bool:
    t = (text or "").upper()
    return "CICLO HURACAN" in t and "NO COD PRODUCTO" in t and ("TOTAL DE LÍNEA" in t or "TOTAL DE LINEA" in t)


def is_brujo_caribeno(text: str) -> bool:
    t = (text or "").upper()
    return "EL BRUJO CARIBEÑO" in t and "CÓDIGO UNIDAD CANTIDAD PRECIO" in t


def is_erial_office_depot(text: str) -> bool:
    t = (text or "").upper()
    return "ERIAL BQ" in t and "LINEA SKU" in t and "IMPUESTO" in t


def is_gustavo_gamboa(text: str) -> bool:
    t = (text or "").upper()
    return "GUSTAVO GAMBOA VILLALOBOS" in t and "# DESCRIPCIÓN / CÓDIGO" in t


# =========================
# HEADER PARSERS
# =========================
def parse_forlan_co_header(text: str, filename: str) -> FinanceInvoice:
    scanned = looks_scanned(text)
    moneda, simbolo = detect_currency(text)

    proveedor = find_first([r"(?m)^(FERRETERIA\s+FORLAN\s+SAS)\s*$"], text)
    proveedor_id = find_first([r"NIT\s*([0-9\.\-]+)"], text)

    cliente = find_first([r"Señores\s+([A-ZÁÉÍÓÚÑ0-9\.\s&\-]+)"], text)
    cliente_nit = find_first([r"Señores.*?\nNIT\s*([0-9\.\-]+)"], text)

    m = re.search(r"No\.\s*([A-Z]{1,5})\s*\n*\s*([0-9]{3,})", text, re.IGNORECASE)
    prefijo = (m.group(1) or "").strip() if m else ""
    consecutivo = (m.group(2) or "").strip() if m else ""
    factura_num = f"{prefijo} {consecutivo}".strip() if prefijo or consecutivo else ""

    fecha = find_first([r"Generaci[oó]n\s*([0-3]\d\/[01]\d\/[12]\d{3},\s*[0-2]\d:[0-5]\d)"], text)
    forma_pago = find_first([r"Forma\s+de\s+pago:\s*\n*([A-Za-zÁÉÍÓÚÑ\s]+)"], text)
    medio_pago = find_first([r"Medio\s+de\s+pago:\s*\n*([A-Za-zÁÉÍÓÚÑ\s\-]+)"], text)

    subtotal_str = find_first([r"Total\s+Bruto\s*([0-9\.,]+)"], text)
    iva_str = find_first([r"IVA\s*19%\s*([0-9\.,]+)"], text)
    total_str = find_first([r"Total\s+a\s+Pagar\s*([0-9\.,]+)"], text)

    oc = find_first([r"Oc:\s*(OC[0-9]+)"], text)
    cufe = find_first([r"CUFE:\s*([a-f0-9]{20,})"], text)
    resol = find_first([r"Autorizaci[oó]n\s+Electr[oó]nica\s+([0-9]+)"], text)
    qr_hint = find_first([r"(CUFE:\s*[a-f0-9]{20,})"], text)

    return FinanceInvoice(
        Documento=filename,
        Pais="CO",
        Tipo_Documento="Factura",
        Proveedor_Razon_Social=proveedor,
        Proveedor_Id_Tributaria=proveedor_id,
        Cliente_Razon_Social=cliente,
        Cliente_Id_Tributaria=cliente_nit,
        Prefijo=prefijo,
        Factura_Numero=factura_num,
        Consecutivo=consecutivo,
        Fecha_Emision=fecha,
        Forma_Pago=forma_pago,
        Medio_Pago=medio_pago,
        Moneda=moneda or "COP",
        Simbolo_Moneda=simbolo or "$",
        Subtotal=parse_number_latam(subtotal_str),
        Impuesto_IVA=parse_number_latam(iva_str),
        Total_Factura=parse_number_latam(total_str),
        OC=oc,
        CUFE=cufe,
        Resolucion_DIAN=resol,
        QR_o_Codigo=qr_hint,
        Probable_Escaneado="SI" if scanned else "NO",
        Metodo_Extraccion="FORLAN CO (header + líneas)",
        Error="",
    )


def parse_navatec_cr_header(text: str, filename: str) -> FinanceInvoice:
    scanned = looks_scanned(text)
    moneda, simbolo = detect_currency(text)

    proveedor = find_first([r"(?m)^(NAVATEC\s+INGENIERIA\s+S\.A\.)\s*$", r"(?m)^(NAVATECO)\s*$"], text)
    ids = re.findall(r"Ident\.\s*Jur[ií]dica:\s*([0-9\-]+)", text, flags=re.IGNORECASE)
    proveedor_id = (ids[0] if len(ids) >= 1 else "").strip()
    cliente_id = (ids[1] if len(ids) >= 2 else "").strip()
    cliente = find_first([r"Receptor\s+([A-ZÁÉÍÓÚÑ0-9\.\s&\-]+)"], text)

    factura_num = find_first([r"Factura\s+Electr[oó]nica\s+N°\s*([0-9]+)"], text)
    fecha = find_first([r"Fecha\s+de\s+Emisi[oó]n:\s*([0-3]\d\/[01]\d\/[12]\d{3}\s+[0-2]\d:[0-5]\d\s*[ap]\.m\.)"], text)
    condicion = find_first([r"Condici[oó]n\s+de\s+venta:\s*([A-Za-zÁÉÍÓÚÑ\s]+)"], text)
    medio = find_first([r"Medio\s+de\s+Pago:\s*([A-Za-zÁÉÍÓÚÑ\s]+)"], text)

    clave = find_first([r"Clave\s+Num[eé]rica:\s*\n*([0-9]{30,})"], text)
    cod_unico = find_first([r"C[oó]digo\s+Único\s+de\s+Consulta:\s*([A-Z0-9]+)"], text)

    subtotal_str = find_first([r"Subtotal\s+Neto\s*¢\s*([0-9\.,]+)"], text)
    iva_str = find_first([r"Total\s+Impuesto\s*¢\s*([0-9\.,]+)"], text)
    total_str = find_first([r"Total\s+Factura:\s*¢\s*([0-9\.,]+)"], text)
    anticipo_str = find_first([r"ANTICIPO\s*¢\s*([0-9\.,]+)"], text)
    saldo_str = find_first([r"SALDO\s*¢\s*([0-9\.,]+)"], text)

    return FinanceInvoice(
        Documento=filename,
        Pais="CR",
        Tipo_Documento="Factura",
        Proveedor_Razon_Social=proveedor,
        Proveedor_Id_Tributaria=proveedor_id,
        Cliente_Razon_Social=cliente,
        Cliente_Id_Tributaria=cliente_id,
        Factura_Numero=factura_num,
        Fecha_Emision=fecha,
        Condicion_Venta=condicion,
        Medio_Pago=medio,
        Moneda=moneda or "CRC",
        Simbolo_Moneda=simbolo or "¢",
        Subtotal=parse_number_latam(subtotal_str),
        Impuesto_IVA=parse_number_latam(iva_str),
        Total_Factura=parse_number_latam(total_str),
        Anticipo=parse_number_latam(anticipo_str),
        Saldo=parse_number_latam(saldo_str),
        Clave_Numerica=clave,
        Codigo_Unico_Consulta=cod_unico,
        Probable_Escaneado="SI" if scanned else "NO",
        Metodo_Extraccion="NAVATEC CR (header + líneas)",
        Error="",
    )


def parse_tribu_hacienda_cr_header(text: str, filename: str) -> FinanceInvoice:
    # Basic header extraction
    scanned = looks_scanned(text)
    moneda, simbolo = detect_currency(text)

    proveedor = find_first([r"Nombre:\s*([A-ZÁÉÍÓÚÑ\s]+)\nNombre comercial:"], text)
    proveedor_id = find_first([r"C[eé]dula:\s*([0-9]+)"], text)
    consecutivo = find_first([r"Consecutivo:\s*([0-9]+)"], text)
    clave = find_first([r"Clave:\s*([0-9]{30,})"], text)
    fecha = find_first([r"Fecha:\s*([0-3]\d\/[01]\d\/[12]\d{3}\s+[0-2]\d:[0-5]\d:[0-5]\d)"], text)

    subtotal_str = find_first([r"Total\s+venta\s+neta\s*([0-9\.,]+)"], text)
    iva_str = find_first([r"Total\s+impuestos\s*([0-9\.,]+)"], text)
    total_str = find_first([r"Total\s+comprobante\s*([0-9\.,]+)"], text)

    return FinanceInvoice(
        Documento=filename,
        Pais="CR",
        Tipo_Documento="Factura",
        Proveedor_Razon_Social=proveedor,
        Proveedor_Id_Tributaria=proveedor_id,
        Cliente_Razon_Social=find_first([r"DATOS\s+DEL\s+CLIENTE\s+Nombre:\s*([A-ZÁÉÍÓÚÑ\s]+)"], text),
        Cliente_Id_Tributaria=find_first([r"DATOS\s+DEL\s+CLIENTE.*?C[eé]dula:\s*([0-9]+)"], text),
        Factura_Numero=consecutivo,
        Consecutivo=consecutivo,
        Fecha_Emision=fecha,
        Condicion_Venta=find_first([r"Condici[oó]n\s+de\s+Venta:\s*([A-Za-zÁÉÍÓÚÑ\s]+)"], text),
        Medio_Pago=find_first([r"Medio\s+de\s+Pago:\s*([A-Za-zÁÉÍÓÚÑ\s\-]+)"], text),
        Moneda=moneda or "CRC",
        Simbolo_Moneda=simbolo or "¢",
        Subtotal=parse_number_latam(subtotal_str),
        Impuesto_IVA=parse_number_latam(iva_str),
        Total_Factura=parse_number_latam(total_str),
        Clave_Numerica=clave,
        Probable_Escaneado="SI" if scanned else "NO",
        Metodo_Extraccion="TRIBU-CR / Hacienda (header + líneas)",
        Error="",
    )


def parse_generic_header(text: str, filename: str, pais_hint: str = "") -> FinanceInvoice:
    scanned = looks_scanned(text)
    moneda, simbolo = detect_currency(text)

    proveedor = find_first(
        [
            r"Raz[oó]n\s+Social[:\s]+([A-ZÁÉÍÓÚÑ0-9&\-\.\s]{4,})",
            r"Nombre:\s*([A-ZÁÉÍÓÚÑ0-9\.\s&\-]+)\n",
            r"Emisor[:\s]+([A-ZÁÉÍÓÚÑ0-9&\-\.\s]{4,})",
        ],
        text,
    ).split("\n")[0].strip()

    proveedor_id = find_first(
        [
            r"NIT[:\s]*([0-9\.\-]{6,20})",
            r"Identificaci[oó]n:\s*([0-9]+)",
            r"C[eé]dula:\s*([0-9]+)",
            r"Ident\.\s*Jur[ií]dica:\s*([0-9\-]+)",
        ],
        text,
    )

    factura_num = find_first(
        [
            r"Factura\s+Electr[oó]nica[:\s#]*([0-9]{8,})",
            r"Consecutivo:\s*([0-9]+)",
            r"Factura\s*(?:No\.|Nro\.|N°|#)?\s*[:\s]*([A-Z0-9\-]{3,})",
        ],
        text,
    )

    fecha = find_first(
        [
            r"Fecha:\s*([0-3]\d\/[01]\d\/[12]\d{3}\s+[0-2]\d:[0-5]\d:[0-5]\d)",
            r"Fecha\s+y\s+Hora\s+de\s+Emisi[oó]n:\s*([0-3]\d\/[01]\d\/[12]\d{3}\s+[0-2]\d:[0-5]\d:[0-5]\d\s*[AP]M)",
            r"Fecha\s+de\s+Emisi[oó]n:\s*([0-3]\d\/[01]\d\/[12]\d{3}\s*[0-2]?\d:[0-5]\d)",
        ],
        text,
    )

    inv = FinanceInvoice(
        Documento=filename,
        Pais=pais_hint,
        Tipo_Documento="Factura",
        Proveedor_Razon_Social=proveedor,
        Proveedor_Id_Tributaria=proveedor_id,
        Factura_Numero=factura_num,
        Fecha_Emision=fecha,
        Moneda=moneda,
        Simbolo_Moneda=simbolo,
        Probable_Escaneado="SI" if scanned else "NO",
        Metodo_Extraccion="Genérico (header)",
        Error="",
    )
    return inv


# =========================
# LINE ITEMS PARSERS
# =========================
def items_forlan_co(text: str, inv: FinanceInvoice) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for ln in lines(text):
        m = re.match(
            r"^(?P<item>\d+)\s+(?P<codigo>\d{3,})\s+(?P<qty>\d+(?:\.\d+)?)\s+(?P<desc>.+?)\s+(?P<unit>[0-9\.,]+)\s+(?P<bruto>[0-9\.,]+)\s+(?P<total>[0-9\.,]+)\s*$",
            ln
        )
        if not m:
            continue
        out.append({
            "Factura_Numero": inv.Factura_Numero,
            "Documento": inv.Documento,
            "Linea": m.group("item"),
            "Codigo_Item": m.group("codigo"),
            "Descripcion": m.group("desc").strip(),
            "Cantidad": parse_number_latam(m.group("qty")),
            "Unidad": "",
            "Precio_Unitario": parse_number_latam(m.group("unit")),
            "Descuento": None,
            "Subtotal_Linea": parse_number_latam(m.group("bruto")),
            "Impuesto_Linea": None,
            "Total_Linea": parse_number_latam(m.group("total")),
            "Moneda": inv.Moneda,
            "Pais": inv.Pais,
            "Marca_Costo": "",
            "Cuenta_Costo": "",
            "Descripcion_Raw": ln,
        })
    return out


def items_navatec_cr(text: str, inv: FinanceInvoice) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    pat = re.compile(
        r"^(?P<linea>\d{3})\s+"
        r"(?P<cantidad>\d+(?:\.\d+)?)\s+"
        r"(?P<unidad>\w+)\s+"
        r"(?P<codigo>[A-Z0-9]+)\s+"
        r"(?P<desc>.+?)\s+"
        r"(?P<precio>[0-9\.,]+)\s+"
        r"(?P<descuento>[0-9\.,]+)\s+"
        r"(?P<subtotal>[0-9\.,]+)\s+"
        r"(?P<imp>[0-9\.,]+)\s*$"
    )
    for ln in lines(text):
        m = pat.match(ln)
        if not m:
            continue

        subtotal_linea = parse_number_latam(m.group("subtotal"))
        imp = parse_number_latam(m.group("imp"))
        total_linea = (subtotal_linea + imp) if (subtotal_linea is not None and imp is not None) else None

        out.append({
            "Factura_Numero": inv.Factura_Numero,
            "Documento": inv.Documento,
            "Linea": m.group("linea"),
            "Codigo_Item": m.group("codigo"),
            "Descripcion": m.group("desc").strip(),
            "Cantidad": parse_number_latam(m.group("cantidad")),
            "Unidad": m.group("unidad"),
            "Precio_Unitario": parse_number_latam(m.group("precio")),
            "Descuento": parse_number_latam(m.group("descuento")),
            "Subtotal_Linea": subtotal_linea,
            "Impuesto_Linea": imp,
            "Total_Linea": total_linea,
            "Moneda": inv.Moneda,
            "Pais": inv.Pais,
            "Marca_Costo": "",
            "Cuenta_Costo": "",
            "Descripcion_Raw": ln,
        })
    return out


def items_tribu_hacienda_cr(text: str, inv: FinanceInvoice) -> List[Dict[str, Any]]:
    """
    TRIBU-CR / Hacienda:
    Busca líneas:
      <linea> <codigo> <desc>
    y luego lee un bloque siguiente con valores.
    """
    out: List[Dict[str, Any]] = []
    ln_list = lines(text)

    i = 0
    while i < len(ln_list):
        ln = ln_list[i]
        m = re.match(r"^(?P<linea>\d+)\s+(?P<codigo>\d{10,})\s+(?P<desc>.+)$", ln)
        if not m:
            i += 1
            continue

        linea = m.group("linea")
        codigo = m.group("codigo")
        desc = m.group("desc").strip()

        j = i + 1
        blob = ""
        while j < len(ln_list):
            if ln_list[j].upper().startswith("OBSERVACIONES"):
                break
            if re.match(r"^\d+\s+\d{10,}\s+", ln_list[j]):  # next item
                break
            blob += " " + ln_list[j]
            j += 1

        # qty
        qty = find_first([r"(\d+,\d+)\s+Unidad", r"(\d+,\d+)\s+Servicios", r"(\d+,\d+)\s+\w+"], blob)
        nums = re.findall(r"[0-9]{1,3}(?:[0-9\.,]*[0-9])", blob)

        precio = monto = descuento = total = None
        if len(nums) >= 4:
            precio = parse_number_latam(nums[-4])
            monto = parse_number_latam(nums[-3])
            descuento = parse_number_latam(nums[-2])
            total = parse_number_latam(nums[-1])

        out.append({
            "Factura_Numero": inv.Factura_Numero,
            "Documento": inv.Documento,
            "Linea": linea,
            "Codigo_Item": codigo,
            "Descripcion": desc,
            "Cantidad": parse_number_latam(qty),
            "Unidad": "Unidad",
            "Precio_Unitario": precio,
            "Descuento": descuento,
            "Subtotal_Linea": monto,
            "Impuesto_Linea": None,
            "Total_Linea": total,
            "Moneda": inv.Moneda,
            "Pais": inv.Pais,
            "Marca_Costo": "",
            "Cuenta_Costo": "",
            "Descripcion_Raw": (ln + " | " + blob.strip())[:1000],
        })

        i = j
    return out


def items_ciclo_huracan(text: str, inv: FinanceInvoice) -> List[Dict[str, Any]]:
    """
    CICLO HURACAN:
      <No> <Cod> <Producto...>
      luego bloque con qty/unid y total
    """
    out: List[Dict[str, Any]] = []
    ln_list = lines(text)

    i = 0
    while i < len(ln_list):
        ln = ln_list[i]
        m = re.match(r"^(?P<linea>\d+)\s+(?P<codigo>\d+)\s+(?P<desc>.+)$", ln)
        if not m:
            i += 1
            continue

        linea = m.group("linea")
        codigo = m.group("codigo")
        desc = m.group("desc").strip()

        blob = ""
        j = i + 1
        while j < len(ln_list):
            if ln_list[j].upper().startswith("COMENTARIO"):
                break
            if re.match(r"^\d+\s+\d+\s+", ln_list[j]):  # next item
                break
            blob += " " + ln_list[j]
            j += 1

        qty = find_first([r"(\d+(?:\.\d+)?)\s+Unid"], blob)
        # find last CRC amount as total
        total = find_first([r"CRC\s*([0-9\.,]+)\s*$"], blob)
        # first CRC amount as price-ish
        precio = find_first([r"CRC\s*([0-9\.,]+)"], blob)

        imp = None
        m_imp = re.search(r"IVA\s*13%.*?CRC\s*([0-9\.,]+)", blob, re.IGNORECASE)
        if m_imp:
            imp = parse_number_latam(m_imp.group(1))

        out.append({
            "Factura_Numero": inv.Factura_Numero,
            "Documento": inv.Documento,
            "Linea": linea,
            "Codigo_Item": codigo,
            "Descripcion": desc,
            "Cantidad": parse_number_latam(qty),
            "Unidad": "Unid",
            "Precio_Unitario": parse_number_latam(precio),
            "Descuento": 0.0,
            "Subtotal_Linea": None,
            "Impuesto_Linea": imp,
            "Total_Linea": parse_number_latam(total),
            "Moneda": inv.Moneda,
            "Pais": inv.Pais,
            "Marca_Costo": "",
            "Cuenta_Costo": "",
            "Descripcion_Raw": (ln + " | " + blob.strip())[:1000],
        })

        i = j
    return out


def items_brujo_caribeno(text: str, inv: FinanceInvoice) -> List[Dict[str, Any]]:
    """
    EL BRUJO CARIBEÑO:
      Descripción larga arriba; luego una línea numérica:
      C01 Al 1.00 300,000.00 0.00 300,000.00 39,000.00
    """
    out: List[Dict[str, Any]] = []
    ln_list = lines(text)

    last_desc = ""
    line_count = 0

    for ln in ln_list:
        # captura descripción tipo servicio
        if "Servicios de alquiler" in ln:
            last_desc = ln.strip()

        m = re.match(
            r"^(?P<codigo>[A-Z0-9]+)\s+(?P<unidad>\w+)\s+(?P<qty>\d+(?:\.\d+)?)\s+(?P<precio>[0-9\.,]+)\s+(?P<descnt>[0-9\.,]+)\s+(?P<subt>[0-9\.,]+)\s+(?P<imp>[0-9\.,]+)\s*$",
            ln
        )
        if not m:
            continue

        line_count += 1
        subtotal = parse_number_latam(m.group("subt"))
        imp = parse_number_latam(m.group("imp"))
        total = (subtotal + imp) if (subtotal is not None and imp is not None) else None

        out.append({
            "Factura_Numero": inv.Factura_Numero,
            "Documento": inv.Documento,
            "Linea": str(line_count),
            "Codigo_Item": m.group("codigo"),
            "Descripcion": last_desc or "SERVICIO",
            "Cantidad": parse_number_latam(m.group("qty")),
            "Unidad": m.group("unidad"),
            "Precio_Unitario": parse_number_latam(m.group("precio")),
            "Descuento": parse_number_latam(m.group("descnt")),
            "Subtotal_Linea": subtotal,
            "Impuesto_Linea": imp,
            "Total_Linea": total,
            "Moneda": inv.Moneda,
            "Pais": inv.Pais,
            "Marca_Costo": "",
            "Cuenta_Costo": "",
            "Descripcion_Raw": ln,
        })

    return out


def items_erial_office_depot(text: str, inv: FinanceInvoice) -> List[Dict[str, Any]]:
    """
    ERIAL BQ (Office Depot):
      1 3212900039900 ... 1.00 Unid 876.11 876.11 113.89 13.00 0.00 990.00
    """
    out: List[Dict[str, Any]] = []
    for ln in lines(text):
        m = re.match(
            r"^(?P<linea>\d+)\s+(?P<sku>\d{10,})\s+(?P<desc>.+?)\s+(?P<qty>\d+(?:\.\d+)?)\s+(?P<uni>\w+)\s+(?P<pu>[0-9\.,]+)\s+(?P<subt>[0-9\.,]+)\s+(?P<imp>[0-9\.,]+)\s+(?P<pct>[0-9\.,]+)\s+(?P<descnt>[0-9\.,]+)\s+(?P<total>[0-9\.,]+)\s*$",
            ln
        )
        if not m:
            continue

        out.append({
            "Factura_Numero": inv.Factura_Numero,
            "Documento": inv.Documento,
            "Linea": m.group("linea"),
            "Codigo_Item": m.group("sku"),
            "Descripcion": m.group("desc").strip(),
            "Cantidad": parse_number_latam(m.group("qty")),
            "Unidad": m.group("uni"),
            "Precio_Unitario": parse_number_latam(m.group("pu")),
            "Descuento": parse_number_latam(m.group("descnt")),
            "Subtotal_Linea": parse_number_latam(m.group("subt")),
            "Impuesto_Linea": parse_number_latam(m.group("imp")),
            "Total_Linea": parse_number_latam(m.group("total")),
            "Moneda": inv.Moneda,
            "Pais": inv.Pais,
            "Marca_Costo": "",
            "Cuenta_Costo": "",
            "Descripcion_Raw": ln,
        })
    return out


def items_gustavo_gamboa(text: str, inv: FinanceInvoice) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for ln in lines(text):
        m = re.match(
            r"^(?P<linea>\d+)\s+(?P<codigo>[A-Z0-9]+)\s+(?P<desc>.+?)\s+(?P<qty>\d+(?:\.\d+)?)\s+(?P<precio>[0-9\.,]+)\s+(?P<uni>Serv\s+Prof|\w+)\s+(?P<descnt>[0-9\.,]+)\s+(?P<pct>[0-9\.,]+)\s+%.*?\s+(?P<imp>[0-9\.,]+)\s+(?P<total>[0-9\.,]+)\s*$",
            ln,
            re.IGNORECASE
        )
        if not m:
            continue

        total = parse_number_latam(m.group("total"))
        imp = parse_number_latam(m.group("imp"))
        subtotal = (total - imp) if (total is not None and imp is not None) else None

        out.append({
            "Factura_Numero": inv.Factura_Numero,
            "Documento": inv.Documento,
            "Linea": m.group("linea"),
            "Codigo_Item": m.group("codigo"),
            "Descripcion": m.group("desc").strip(),
            "Cantidad": parse_number_latam(m.group("qty")),
            "Unidad": m.group("uni").strip(),
            "Precio_Unitario": parse_number_latam(m.group("precio")),
            "Descuento": parse_number_latam(m.group("descnt")),
            "Subtotal_Linea": subtotal,
            "Impuesto_Linea": imp,
            "Total_Linea": total,
            "Moneda": inv.Moneda,
            "Pais": inv.Pais,
            "Marca_Costo": "",
            "Cuenta_Costo": "",
            "Descripcion_Raw": ln,
        })
    return out


# =========================
# EXCEL FORMATTING + GROUPING
# =========================
def autosize_columns(ws):
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in col_cells:
            val = "" if cell.value is None else str(cell.value)
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)


def apply_global_excel_formatting(wb):
    font = Font(name="Century Gothic", size=10)
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for row in ws.iter_rows():
            for cell in row:
                cell.font = font

        autosize_columns(ws)


def group_line_items_by_invoice(ws, factura_col_letter: str = "A"):
    ws.sheet_properties.outlinePr.summaryBelow = True
    ws.sheet_view.showOutlineSymbols = True

    current = None
    start_row = None
    max_row = ws.max_row

    for r in range(2, max_row + 1):
        val = ws[f"{factura_col_letter}{r}"].value
        if val != current:
            if current is not None and start_row is not None:
                end_row = r - 1
                if end_row > start_row:
                    ws.row_dimensions.group(start_row, end_row, outline_level=1, hidden=False)
            current = val
            start_row = r

    if current is not None and start_row is not None:
        end_row = max_row
        if end_row > start_row:
            ws.row_dimensions.group(start_row, end_row, outline_level=1, hidden=False)


def build_excel_bytes(df_fin: pd.DataFrame, df_lines: pd.DataFrame, df_audit: pd.DataFrame) -> bytes:
    out = io.BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        df_fin.to_excel(writer, index=False, sheet_name="FINANZAS_FACTURAS")
        df_lines.to_excel(writer, index=False, sheet_name="LINEAS_FACTURA")
        df_audit.to_excel(writer, index=False, sheet_name="AUDITORIA_TEXTO")

    out.seek(0)
    wb = load_workbook(out)

    apply_global_excel_formatting(wb)

    if "LINEAS_FACTURA" in wb.sheetnames:
        ws = wb["LINEAS_FACTURA"]
        group_line_items_by_invoice(ws, factura_col_letter="A")

    final = io.BytesIO()
    wb.save(final)
    return final.getvalue()


# =========================
# CORE PROCESSING
# =========================
def process_files(files, include_audit: bool, audit_chars: int) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    fin_rows: List[Dict[str, Any]] = []
    line_rows: List[Dict[str, Any]] = []
    audit_rows: List[Dict[str, Any]] = []

    for uf in files:
        pdf_bytes = uf.read()
        try:
            text = extract_text_pypdf(pdf_bytes)

            # Default currency
            cur, sym = detect_currency(text)

            # Header + Lines by type
            if is_forlan_co(text):
                inv = parse_forlan_co_header(text, uf.name)
                items = items_forlan_co(text, inv)

            elif is_navatec_cr(text):
                inv = parse_navatec_cr_header(text, uf.name)
                items = items_navatec_cr(text, inv)

            elif is_tribu_cr_hacienda(text):
                inv = parse_tribu_hacienda_cr_header(text, uf.name)
                items = items_tribu_hacienda_cr(text, inv)

            elif is_ciclo_huracan(text):
                inv = parse_generic_header(text, uf.name, pais_hint="CR")
                inv.Pais = "CR"
                if not inv.Moneda:
                    inv.Moneda, inv.Simbolo_Moneda = (cur or "CRC"), (sym or "¢")
                inv.Metodo_Extraccion = "CICLO HURACAN (header + líneas)"
                items = items_ciclo_huracan(text, inv)

            elif is_brujo_caribeno(text):
                inv = parse_generic_header(text, uf.name, pais_hint="CR")
                inv.Pais = "CR"
                if not inv.Moneda:
                    inv.Moneda, inv.Simbolo_Moneda = (cur or "CRC"), (sym or "¢")
                inv.Metodo_Extraccion = "EL BRUJO CARIBEÑO (header + líneas)"
                items = items_brujo_caribeno(text, inv)

            elif is_erial_office_depot(text):
                inv = parse_generic_header(text, uf.name, pais_hint="CR")
                inv.Pais = "CR"
                if not inv.Moneda:
                    inv.Moneda, inv.Simbolo_Moneda = (cur or "CRC"), (sym or "¢")
                inv.Metodo_Extraccion = "ERIAL BQ (header + líneas)"
                items = items_erial_office_depot(text, inv)

            elif is_gustavo_gamboa(text):
                inv = parse_generic_header(text, uf.name, pais_hint="CR")
                inv.Pais = "CR"
                if not inv.Moneda:
                    inv.Moneda, inv.Simbolo_Moneda = (cur or "CRC"), (sym or "¢")
                inv.Metodo_Extraccion = "GUSTAVO GAMBOA (header + líneas)"
                items = items_gustavo_gamboa(text, inv)

            else:
                inv = parse_generic_header(text, uf.name)
                if not inv.Moneda:
                    inv.Moneda, inv.Simbolo_Moneda = cur, sym
                inv.Metodo_Extraccion = inv.Metodo_Extraccion or "Genérico (header)"
                items = []

            fin_row = inv.__dict__
            fin_rows.append({c: fin_row.get(c, "") for c in FIN_COLS})

            # Lines (if none -> single marker row)
            if items:
                for it in items:
                    line_rows.append({c: it.get(c, "") for c in LINE_COLS})
            else:
                line_rows.append({
                    "Factura_Numero": inv.Factura_Numero,
                    "Documento": inv.Documento,
                    "Linea": "",
                    "Codigo_Item": "",
                    "Descripcion": "",
                    "Cantidad": None,
                    "Unidad": "",
                    "Precio_Unitario": None,
                    "Descuento": None,
                    "Subtotal_Linea": None,
                    "Impuesto_Linea": None,
                    "Total_Linea": None,
                    "Moneda": inv.Moneda,
                    "Pais": inv.Pais,
                    "Marca_Costo": "",
                    "Cuenta_Costo": "",
                    "Descripcion_Raw": "SIN_LINEAS_DETECTADAS",
                })

            if include_audit:
                audit_rows.append({
                    "Documento": uf.name,
                    "Longitud_Texto": len(text),
                    "Texto": (text or "")[:audit_chars],
                })

        except Exception as e:
            err_row = {c: "" for c in FIN_COLS}
            err_row["Documento"] = uf.name
            err_row["Metodo_Extraccion"] = "ERROR"
            err_row["Error"] = str(e)
            fin_rows.append(err_row)

            line_rows.append({
                "Factura_Numero": "",
                "Documento": uf.name,
                "Linea": "",
                "Codigo_Item": "",
                "Descripcion": "",
                "Cantidad": None,
                "Unidad": "",
                "Precio_Unitario": None,
                "Descuento": None,
                "Subtotal_Linea": None,
                "Impuesto_Linea": None,
                "Total_Linea": None,
                "Moneda": "",
                "Pais": "",
                "Marca_Costo": "",
                "Cuenta_Costo": "",
                "Descripcion_Raw": f"ERROR: {e}",
            })

            if include_audit:
                audit_rows.append({"Documento": uf.name, "Longitud_Texto": 0, "Texto": f"ERROR: {e}"})

    df_fin = pd.DataFrame(fin_rows, columns=FIN_COLS)
    df_lines = pd.DataFrame(line_rows, columns=LINE_COLS)
    df_audit = pd.DataFrame(audit_rows) if include_audit else pd.DataFrame(columns=["Documento", "Longitud_Texto", "Texto"])

    df_lines["Factura_Numero"] = df_lines["Factura_Numero"].fillna("")
    df_lines = df_lines.sort_values(by=["Factura_Numero", "Documento", "Linea"], kind="stable").reset_index(drop=True)

    return df_fin, df_lines, df_audit


# =========================
# UI: UPLOAD + CONTROLS
# =========================
with st.container():
    c1, c2, c3 = st.columns([2, 1, 1], vertical_alignment="bottom")

    with c1:
        uploaded_files = st.file_uploader(
            "Sube uno o varios PDFs (facturas)",
            type=["pdf"],
            accept_multiple_files=True,
        )
        if uploaded_files:
            st.markdown('<div class="file-scroll">', unsafe_allow_html=True)
            for f in uploaded_files:
                st.write(f"📄 {f.name} — {round(f.size/1024, 1)} KB")
            st.markdown("</div>", unsafe_allow_html=True)

    with c2:
        show_audit = st.checkbox("Mostrar auditoría", value=False)
        show_text = st.checkbox("Mostrar texto", value=False, help="Solo en detalle de la factura seleccionada.")
        audit_chars = st.slider("Texto auditoría (caracteres)", 2000, 32000, 12000, 500)

    with c3:
        process_btn = st.button("🚀 Procesar", type="primary", use_container_width=True)

st.divider()


# =========================
# RUN PROCESS
# =========================
if process_btn:
    if not uploaded_files:
        st.error("Sube al menos un PDF.")
        st.stop()

    with st.spinner("Procesando facturas…"):
        df_fin, df_lines, df_audit = process_files(uploaded_files, include_audit=show_audit, audit_chars=audit_chars)

    st.session_state["df_fin"] = df_fin
    st.session_state["df_lines"] = df_lines
    st.session_state["df_audit"] = df_audit

    st.success(f"✅ Proceso completado: {len(df_fin)} factura(s).")


# =========================
# REPORT (Unified) + ZOOM DETAIL
# =========================
if "df_fin" in st.session_state and not st.session_state["df_fin"].empty:
    df_fin = st.session_state["df_fin"]
    df_lines = st.session_state.get("df_lines", pd.DataFrame())
    df_audit = st.session_state.get("df_audit", pd.DataFrame())

    # Filters
    f1, f2, f3 = st.columns([1.5, 1, 1])
    with f1:
        search = st.text_input("🔎 Buscar (Proveedor, NIT, Factura, Documento)", value="")
    with f2:
        paises = sorted([p for p in df_fin.get("Pais", pd.Series(dtype=str)).dropna().unique().tolist() if str(p).strip()])
        pais_filter = st.multiselect("🌎 País", options=paises, default=paises)
    with f3:
        metodo_vals = sorted([m for m in df_fin.get("Metodo_Extraccion", pd.Series(dtype=str)).dropna().unique().tolist() if str(m).strip()])
        metodo_filter = st.multiselect("🧠 Método", options=metodo_vals, default=metodo_vals)

    view = df_fin.copy()
    if pais_filter and "Pais" in view.columns:
        view = view[view["Pais"].isin(pais_filter)]
    if metodo_filter and "Metodo_Extraccion" in view.columns:
        view = view[view["Metodo_Extraccion"].isin(metodo_filter)]

    if search.strip():
        s = search.strip().lower()
        cols = [c for c in ["Proveedor_Razon_Social", "Proveedor_Id_Tributaria", "Factura_Numero", "Documento", "Cliente_Razon_Social"] if c in view.columns]
        if cols:
            mask = False
            for c in cols:
                mask = mask | view[c].astype(str).str.lower().str.contains(s, na=False)
            view = view[mask]

    left, right = st.columns([1.1, 1.4], gap="large")

    # LEFT
    with left:
        st.subheader("📌 Facturas (Reporte)")
        st.caption("Usa búsqueda/filtros. Selecciona una factura para ver detalle.")

        display_cols = [
            "Documento", "Pais", "Proveedor_Razon_Social", "Proveedor_Id_Tributaria",
            "Factura_Numero", "Fecha_Emision", "Total_Factura", "Moneda", "Metodo_Extraccion"
        ]
        view_disp = view[[c for c in display_cols if c in view.columns]].reset_index(drop=True)

        if view_disp.empty:
            st.info("No hay resultados con los filtros actuales.")
            st.stop()

        options = []
        for i, r in view_disp.iterrows():
            factura = r.get("Factura_Numero", "")
            prov = (r.get("Proveedor_Razon_Social", "") or "")[:28]
            total = r.get("Total_Factura", "")
            options.append(f"{i+1:03d} | {factura} | {prov} | {total}")

        selected = st.selectbox("📄 Selección", options=options, index=0)
        sel_idx = int(selected.split("|")[0].strip()) - 1
        selected_row = view_disp.iloc[sel_idx].to_dict()

        st.markdown('<div class="table-scroll">', unsafe_allow_html=True)
        st.dataframe(view_disp, use_container_width=True, height=520)
        st.markdown("</div>", unsafe_allow_html=True)

    # RIGHT
    with right:
        st.subheader("🔍 Detalle de la factura")
        st.caption("Encabezado + Líneas + Auditoría (opcional).")

        doc = selected_row.get("Documento", "")
        facnum = selected_row.get("Factura_Numero", "")

        tabs = st.tabs(["Encabezado", "Líneas", "Auditoría"])

        with tabs[0]:
            header_df = df_fin[df_fin["Documento"] == doc].copy() if "Documento" in df_fin.columns else df_fin.copy()
            st.dataframe(header_df, use_container_width=True)

        with tabs[1]:
            if df_lines is None or df_lines.empty:
                st.info("No hay líneas detectadas.")
            else:
                lines_df = df_lines.copy()
                if "Documento" in lines_df.columns:
                    lines_df = lines_df[lines_df["Documento"] == doc]
                elif "Factura_Numero" in lines_df.columns and facnum:
                    lines_df = lines_df[lines_df["Factura_Numero"] == facnum]

                st.markdown('<div class="table-scroll">', unsafe_allow_html=True)
                st.dataframe(lines_df, use_container_width=True, height=520)
                st.markdown("</div>", unsafe_allow_html=True)

        with tabs[2]:
            if not show_audit:
                st.info("Activa 'Mostrar auditoría' para ver esta sección.")
            else:
                if df_audit is None or df_audit.empty:
                    st.info("No hay auditoría disponible.")
                else:
                    aud = df_audit[df_audit["Documento"] == doc].copy() if "Documento" in df_audit.columns else df_audit.copy()
                    st.dataframe(aud, use_container_width=True)

                    if show_text and not aud.empty and "Texto" in aud.columns:
                        st.text_area("Texto extraído (preview)", value=str(aud.iloc[0]["Texto"]), height=320)

    st.divider()

    st.subheader("⬇️ Exportar")
    st.caption("Descarga Excel final con agrupación por factura y formato contable.")

    excel_bytes = build_excel_bytes(df_fin, df_lines, df_audit)
    filename = f"SED_Facturas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    st.download_button(
        "📥 Descargar Excel",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
